import os
import openpyxl
from define import STRING_EMPTY, excel_column, excel_coordinate
import collections


class ExcelFiledNameDescript(object):
    def __init__(self, **kwargs):
        self.field_desc = None
        self.orignal_str = STRING_EMPTY
        self.name = STRING_EMPTY
        return super().__init__(**kwargs)

    def init(self, field_desc, name_str):
        self.field_desc = field_desc
        self.orignal_str = name_str
        self.name = name_str
        return True


class ExcelFieldTypeDescript(object):
    def __init__(self, **kwargs):
        self.field_desc = None
        self.orignal_str = STRING_EMPTY
        self.type = STRING_EMPTY
        return super().__init__(**kwargs)

    def init(self, field_desc, type_str):
        self.field_desc = field_desc
        self.orignal_str = type_str
        return self.parse()

    def parse(self):
        segment_strs = self.orignal_str.strip().split(';')
        if len(segment_strs) <= 0 or not segment_strs[0]:
            return False
        if not self.parse_type(segment_strs[0]):
            return False
        return True
    
    def parse_type(self, type_str):
        self.type = type_str
        return True


class ExcelFieldDescript(object):
    def __init__(self, **kwargs):
        self.excel_desc = None
        self.column = ""
        self.name_desc = None
        self.type_desc = None
        return super().__init__(**kwargs)

    def init(self, excel_desc, column, name_str, type_str):
        self.column = column
        self.name_desc = ExcelFiledNameDescript()
        self.type_desc = ExcelFieldTypeDescript()
        all_ok = True
        all_ok = all_ok and self.name_desc.init(self, name_str)
        all_ok = all_ok and self.type_desc.init(self, type_str)
        return all_ok

class ExcelExtraFieldDescript(object):
    def __init__(self, **kwargs):
        self.field_type = []
        field_name = STRING_EMPTY
        return super().__init__(**kwargs)

class ExcelDescript(object):
    START_ROW_CELL = "A1"
    EXTRA_FIELD_ROW = 2

    def __init__(self, **kwargs):
        self.start_row = -1
        self.min_column = -1
        self.max_column = -2
        self.field_descs = []
        self.sheet_data = None
        self.extra_field_descs = []
        return super().__init__(**kwargs)

    def init(self, sheet_data):
        self.sheet_data = sheet_data
        self.start_row = int(sheet_data[ExcelDescript.START_ROW_CELL].value)
        if self.start_row <= 0: 
            return False
        all_ok = True
        for cell in self.sheet_data[ExcelDescript.EXTRA_FIELD_ROW:ExcelDescript.EXTRA_FIELD_ROW]:
            if not cell.value or not cell.value.strip(): 
                continue
            (extra_field_type, extra_field_name) = cell.value.split(':')[0:2]
            if not extra_field_type or not extra_field_name:
                all_ok = False
                break
            extra_field_desc = ExcelExtraFieldDescript()
            extra_field_desc.field_type = extra_field_type.strip('.').split('.')
            extra_field_desc.field_name = extra_field_name
            self.extra_field_descs.append(extra_field_desc)
        if all_ok:
            for name_cell, type_cell in self.sheet_data.iter_cols(min_row=self.start_row, max_row=self.start_row+1):
                if not name_cell.value or not type_cell.value:
                    if self.min_column < 0: continue
                    else: break
                if self.min_column < 0: 
                    self.min_column = name_cell.col_idx
                self.max_column = name_cell.col_idx
                field_desc = ExcelFieldDescript()
                if field_desc.init(self, name_cell.col_idx, name_cell.value, type_cell.value):
                    self.field_descs.append(field_desc)
                else: 
                    all_ok = False
                    break
        return all_ok and len(self.field_descs) > 0
            
    @staticmethod
    def load(file_path, sheet_name):
        excel_desc = None
        if os.path.exists(file_path) and os.path.isfile(file_path):
            xls_data = openpyxl.load_workbook(file_path)
            if xls_data and sheet_name in xls_data.sheetnames:
                excel_desc = ExcelDescript()
                if not excel_desc.init(xls_data[sheet_name]):
                    excel_desc = None
        return excel_desc