import os
import typing

import logbook
import openpyxl
import enum
import re

from .sheet_field_type import SheetFieldType

class SheetField(object):
    field_name: str
    field_note: str
    is_key: bool
    col_idx: int
    field_type: SheetFieldType
    field_key_type: SheetFieldType
    field_value_type: SheetFieldType

    def __init__(self):
        self.field_name = None
        self.field_type = None
        self.field_note = None
        self.is_key = False
        self.col_idx = -1

class SheetOpt(object):
    opt_key: str
    opt_value:str

    Logic_Config_Name = "logic_config_name"
    Logic_Config_Fun_Name = "logic_config_fun_name"
    Logic_Config_Struct = "logic_config_struct"

    def __init__(self):
        self.opt_key = None
        self.opt_value = None

    @staticmethod
    def parse(in_str):
        if not in_str:
            return None
        m_ret = re.match(r"__opt__:(.*):(.*)", in_str)
        if not m_ret:
            return None
        ret = SheetOpt()
        ret.opt_key = m_ret.group(1).strip()
        ret.opt_value = m_ret.group(2).strip()
        return ret

class SheetDepict(object):
    Start_Row_Cell = "A1"
    Opt_Start_Row = 2

    field_list: typing.List[SheetField]
    opt_map: typing.Dict[str, SheetOpt]
    work_sheet: openpyxl.workbook.workbook.Worksheet
    fo_key_field: SheetField
    data_list: typing.List[typing.List[typing.Any]]

    def __init__(self):
        self.start_row = -1
        self.min_column = -1
        self.max_column = -2
        self.field_list = []
        self.opt_map = {}
        self.work_sheet = None
        self.data_list = []
        self.fo_key_field = None

    def parse(self, work_sheet: openpyxl.workbook.workbook.Worksheet):
        self.work_sheet = work_sheet
        self.start_row = int(self.work_sheet[SheetDepict.Start_Row_Cell].value)
        # logbook.debug("start row {}", self.start_row)
        for row in self.work_sheet.iter_rows(min_row=2, max_row=self.start_row-1):
            for cell in row:
                if cell.value is not None:
                    # logbook.debug("cell {} {}", type(cell), cell.value)
                    opt: SheetOpt = SheetOpt.parse(cell.value)
                    if opt:
                        self.opt_map[opt.opt_key] = opt
                        # logbook.debug("xxxxxx {} {}", opt.opt_key, opt.opt_value)
        for note_cell, name_cell, type_cell in self.work_sheet.iter_cols(min_row=self.start_row, max_row=self.start_row + 2):
            if not name_cell.value or not type_cell.value:
                if self.min_column < 0:
                    continue
                else:
                    break
            if self.min_column < 0:
                self.min_column = name_cell.col_idx
            self.max_column = name_cell.col_idx
            sheet_field = SheetField()
            sheet_field.col_idx = name_cell.col_idx
            sheet_field.field_note = note_cell.value
            sheet_field.field_name = name_cell.value
            sheet_field.field_type = SheetFieldType()
            str_list = str.split(type_cell.value, ";")
            field_type_str = str_list[0]
            if len(str_list) <= 0 or len(str_list[0]) <= 0:
                return False
            ret, sheet_field.field_type, sheet_field.field_key_type, sheet_field.field_value_type = SheetFieldType.parse_type(str_list[0])
            if not ret:
                return False
            if "config_id" == name_cell.value:
                sheet_field.is_key = True
            self.field_list.append(sheet_field)
        key_field: SheetField = None
        for item in self.field_list:
            if item.is_key:
                if key_field is not None:
                    logbook.error("two key field {} and {}", item.col_idx, key_field.col_idx)
                    return False
                else:
                    key_field = item
        if not key_field:
            logbook.error("no key field")
            return False
        self.fo_key_field = key_field
        if not SheetFieldType.is_base_type(key_field.field_type):
            logbook.error("key field value need to be base type")
            return False
        for row in self.work_sheet.iter_rows(min_row=self.start_row+3):
            if not row[key_field.col_idx - 1].value:
                continue
            data = []
            self.data_list.append(data)
            for field in self.field_list:
                cell_str = row[field.col_idx - 1].value
                field_val = SheetFieldType.extract_data(cell_str, field.field_type, field.field_key_type, field.field_value_type)
                data.append(field_val)
        # logbook.debug("data list {}", self.data_list)
        return  True

def parse_sheet_depict(file_path:str, sheet_name:str)->SheetDepict:
    ret = None
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
        if wb and sheet_name in wb.sheetnames:
            ret = SheetDepict()
            if not ret.parse(wb[sheet_name]):
                ret = None
    return ret


